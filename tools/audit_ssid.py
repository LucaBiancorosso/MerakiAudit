from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config.settings import OUTPUT_DIR
from lib.meraki_client import get_dashboard
from lib.output import write_csv, write_json
from lib.standards import (
    NOT_DEFINED,
    compare_field,
    load_ignored_fields,
    load_standards,
)



# ---------------------------------------------------------------------------
# Unconfigured SSID filter
# ---------------------------------------------------------------------------

import re as _re
_UNCONFIGURED_RE = _re.compile(r'^unconfigured\s+ssid\s+\d+$', _re.IGNORECASE)

def _is_unconfigured(name: str) -> bool:
    """Return True for Meraki placeholder SSIDs like 'Unconfigured SSID 1'."""
    return bool(_UNCONFIGURED_RE.match((name or "").strip()))


# ---------------------------------------------------------------------------
# Colours for the XLSX report
# ---------------------------------------------------------------------------

_PASS_FILL   = PatternFill("solid", fgColor="E2EFDA")  # green
_FAIL_FILL   = PatternFill("solid", fgColor="FCE4D6")  # red/orange
_ND_FILL     = PatternFill("solid", fgColor="F2F2F2")  # grey
_WARN_FILL   = PatternFill("solid", fgColor="FFF2CC")  # yellow
_HDR_FILL    = PatternFill("solid", fgColor="1F4E79")  # dark blue
_HDR_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_TOTAL_FILL  = PatternFill("solid", fgColor="FFF4CC")  # light yellow
_THIN        = Side(style="thin", color="D9D9D9")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")

_RESULT_FILL = {
    "PASS":          _PASS_FILL,
    "FAIL":          _FAIL_FILL,
    NOT_DEFINED:     _ND_FILL,
    "NOT_DEFINED":   _ND_FILL,
    "MISSING_SSID":  _WARN_FILL,
    "NON_STANDARD":  _WARN_FILL,
    "IGNORED":       _ND_FILL,
}


# ---------------------------------------------------------------------------
# Network lookup — same pattern as other tools
# ---------------------------------------------------------------------------

def build_network_lookup(dashboard, org_id: str) -> dict[str, str]:
    networks = dashboard.organizations.getOrganizationNetworks(
        org_id, total_pages="all"
    )
    lookup: dict[str, str] = {}
    for net in networks:
        if not isinstance(net, dict):
            continue
        net_id = net.get("id")
        net_name = net.get("name")
        if net_id:
            lookup[str(net_id).strip()] = net_name
    return lookup


# ---------------------------------------------------------------------------
# SSID fetch
# ---------------------------------------------------------------------------

def fetch_ssids_for_network(dashboard, network_id: str) -> list[dict]:
    """Return the raw SSID list for a network, normalising radius hosts."""
    try:
        ssids = dashboard.wireless.getNetworkWirelessSsids(network_id)
    except Exception as exc:
        print(f"  WARNING: could not fetch SSIDs for {network_id}: {exc}")
        return []

    result = []
    for ssid in ssids:
        # Skip placeholder SSIDs Meraki creates for every unused slot
        if _is_unconfigured(ssid.get("name", "")):
            continue
        radius_servers = ssid.get("radiusServers") or []
        radius_hosts = ",".join(
            srv.get("host", "")
            for srv in radius_servers
            if isinstance(srv, dict)
        )
        result.append(
            {
                "ssid_number":             ssid.get("number"),
                "ssid_name":               ssid.get("name"),
                "enabled":                 str(ssid.get("enabled")),
                "authMode":                ssid.get("authMode"),
                "encryptionMode":          ssid.get("encryptionMode"),
                "wpaEncryptionMode":       ssid.get("wpaEncryptionMode"),
                "ipAssignmentMode":        ssid.get("ipAssignmentMode"),
                "bandSelection":           ssid.get("bandSelection"),
                "minBitrate":              ssid.get("minBitrate"),
                "ssidAdminAccessible":     str(ssid.get("ssidAdminAccessible")),
                "radiusEnabled":           str(ssid.get("radiusEnabled")),
                "radiusAccountingEnabled": str(ssid.get("radiusAccountingEnabled")),
                "radiusHosts":             radius_hosts,
                "splashPage":              ssid.get("splashPage"),
                "walledGardenEnabled":     str(ssid.get("walledGardenEnabled")),
                "visible":                 str(ssid.get("visible")),
            }
        )
    return result


# ---------------------------------------------------------------------------
# Core audit logic
# ---------------------------------------------------------------------------

_AUDITED_FIELDS = [
    "ssid_number", "enabled", "authMode", "encryptionMode",
    "wpaEncryptionMode", "ipAssignmentMode", "bandSelection", "minBitrate",
    "ssidAdminAccessible", "radiusEnabled", "radiusAccountingEnabled",
    "radiusHosts", "splashPage", "walledGardenEnabled", "visible",
]


def audit_network(
    org_id: str,
    org_name: str,
    network_id: str,
    network_name: str,
    live_ssids: list[dict],
    standards: dict[tuple[str, str], dict],
    ignored_fields: set[str],
    single_network_mode: bool = False,
) -> list[dict]:
    """
    Compare live SSIDs against standards for one network.
    Returns a flat list of detail rows, one per (ssid, field).

    single_network_mode: when True (--network-id filter used), suppress
    MISSING_SSID checks — standard SSIDs may simply live in other networks.
    """
    detail_rows: list[dict] = []

    live_by_name = {s["ssid_name"]: s for s in live_ssids}
    live_by_name_lower = {k.lower(): v for k, v in live_by_name.items()}
    # standard keys use lowercase ssid_name (set in load_standards)
    standard_names_lower = {
        ssid_name_lower
        for (oid, ssid_name_lower) in standards
        if oid == org_id.strip().lower()
    }

    # ── Check every live SSID ────────────────────────────────────────────────
    for ssid_name, live in live_by_name.items():
        key = (org_id.strip().lower(), ssid_name.lower())
        standard = standards.get(key)

        if standard is None:
            # SSID exists live but not in standard → NON_STANDARD
            # Still emit one row per field with actual values so the report
            # shows what the rogue SSID is actually configured with.
            for field in _AUDITED_FIELDS:
                detail_rows.append(
                    _make_row(
                        org_id, org_name, network_id, network_name,
                        ssid_name, live.get("ssid_number"),
                        field=field,
                        expected="",
                        actual=str(live.get(field, "")),
                        result="NON_STANDARD",
                    )
                )
            continue

        for field in _AUDITED_FIELDS:
            if field in ignored_fields:
                continue
            expected = standard.get(field, NOT_DEFINED)
            actual   = live.get(field, "")
            check    = compare_field(field, expected, actual, ignored_fields)
            detail_rows.append(
                _make_row(
                    org_id, org_name, network_id, network_name,
                    ssid_name, live.get("ssid_number"),
                    field=field,
                    expected=check["expected"],
                    actual=check["actual"],
                    result=check["result"],
                )
            )

    # ── Check for SSIDs defined in standard but missing from network ─────────
    # Skip when filtering to a single network: the SSID may simply live elsewhere.
    if single_network_mode:
        return detail_rows

    for ssid_name_lower in standard_names_lower:
        if ssid_name_lower not in live_by_name_lower:
            detail_rows.append(
                _make_row(
                    org_id, org_name, network_id, network_name,
                    ssid_name_lower, None,
                    field="—",
                    expected="",
                    actual="",
                    result="MISSING_SSID",
                )
            )

    return detail_rows


def _make_row(
    org_id, org_name, network_id, network_name,
    ssid_name, ssid_number,
    field, expected, actual, result,
) -> dict:
    return {
        "org_id":       org_id,
        "org_name":     org_name,
        "network_id":   network_id,
        "network_name": network_name,
        "ssid_name":    ssid_name,
        "ssid_number":  ssid_number if ssid_number is not None else "",
        "field":        field,
        "expected":     expected,
        "actual":       actual,
        "result":       result,
    }


# ---------------------------------------------------------------------------
# Summary builder
# ---------------------------------------------------------------------------

def build_summary(detail_rows: list[dict]) -> list[dict]:
    """
    Aggregate detail rows into one summary row per (network, ssid).
    Counts: tot_checks, tot_pass, tot_fail, tot_not_defined, tot_ignored,
            non_standard, missing_ssid, overall.
    """
    buckets: dict[tuple, dict] = {}

    for row in detail_rows:
        key = (
            row["org_id"], row["org_name"],
            row["network_id"], row["network_name"],
            row["ssid_name"], row["ssid_number"],
        )
        if key not in buckets:
            buckets[key] = {
                "org_id":       row["org_id"],
                "org_name":     row["org_name"],
                "network_id":   row["network_id"],
                "network_name": row["network_name"],
                "ssid_name":    row["ssid_name"],
                "ssid_number":  row["ssid_number"],
                "tot_checks":    0,
                "tot_pass":      0,
                "tot_fail":      0,
                "tot_not_defined": 0,
                "tot_ignored":   0,
                "non_standard":  False,
                "missing_ssid":  False,
            }

        result = row["result"]
        b = buckets[key]

        if result == "PASS":
            b["tot_checks"] += 1
            b["tot_pass"]   += 1
        elif result == "FAIL":
            b["tot_checks"] += 1
            b["tot_fail"]   += 1
        elif result in (NOT_DEFINED, "NOT_DEFINED"):
            b["tot_not_defined"] += 1
        elif result == "IGNORED":
            b["tot_ignored"] += 1
        elif result == "NON_STANDARD":
            b["non_standard"] = True
        elif result == "MISSING_SSID":
            b["missing_ssid"] = True

    summary_rows = []
    for b in buckets.values():
        if b["non_standard"]:
            overall = "NON_STANDARD"
        elif b["missing_ssid"]:
            overall = "MISSING_SSID"
        elif b["tot_fail"] > 0:
            overall = "FAIL"
        elif b["tot_pass"] > 0:
            overall = "PASS"
        else:
            overall = "NOT_DEFINED"
        b["overall"] = overall
        summary_rows.append(b)

    return summary_rows


# ---------------------------------------------------------------------------
# XLSX report writer
# ---------------------------------------------------------------------------

def _style_header_row(ws) -> None:
    for cell in ws[1]:
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
    ws.row_dimensions[1].height = 25


def _add_table(ws, table_name: str) -> None:
    if ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True, showFirstColumn=False,
        showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(tbl)


def _autowidth(ws, min_w=10, max_w=40) -> None:
    for col in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            min(max(length + 2, min_w), max_w)
        )


def write_xlsx_report(
    detail_rows: list[dict],
    summary_rows: list[dict],
    output_path: Path,
) -> None:
    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.freeze_panes = "A2"

    sum_headers = [
        "org_id", "org_name", "network_id", "network_name",
        "ssid_name", "ssid_number",
        "tot_checks", "tot_pass", "tot_fail", "tot_not_defined",
        "tot_ignored", "non_standard", "missing_ssid", "overall",
    ]
    ws_sum.append(sum_headers)
    _style_header_row(ws_sum)

    for row in sorted(
        summary_rows,
        key=lambda r: (r["org_name"], r["network_name"], r["ssid_name"]),
    ):
        ws_sum.append([row.get(h, "") for h in sum_headers])
        result_cell = ws_sum.cell(row=ws_sum.max_row, column=sum_headers.index("overall") + 1)
        result_cell.fill = _RESULT_FILL.get(str(result_cell.value), _ND_FILL)
        for cell in ws_sum[ws_sum.max_row]:
            cell.border    = _BORDER
            cell.alignment = _LEFT
            cell.font      = Font(name="Arial", size=10)

    _autowidth(ws_sum)
    _add_table(ws_sum, "AuditSummary")

    # ── Detail sheet ─────────────────────────────────────────────────────────
    ws_det = wb.create_sheet("Detail")
    ws_det.freeze_panes = "A2"

    det_headers = [
        "org_id", "org_name", "network_id", "network_name",
        "ssid_name", "ssid_number", "field", "expected", "actual", "result",
    ]
    ws_det.append(det_headers)
    _style_header_row(ws_det)

    for row in sorted(
        detail_rows,
        key=lambda r: (r["org_name"], r["network_name"], r["ssid_name"], r["field"]),
    ):
        ws_det.append([row.get(h, "") for h in det_headers])
        result_cell = ws_det.cell(row=ws_det.max_row, column=det_headers.index("result") + 1)
        result_cell.fill = _RESULT_FILL.get(str(result_cell.value), _ND_FILL)
        for cell in ws_det[ws_det.max_row]:
            cell.border    = _BORDER
            cell.alignment = _LEFT
            cell.font      = Font(name="Arial", size=10)

    _autowidth(ws_det)
    _add_table(ws_det, "AuditDetail")

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

_DETAIL_CSV_FIELDS = [
    "org_id", "org_name", "network_id", "network_name",
    "ssid_name", "ssid_number", "field", "expected", "actual", "result",
]

_SUMMARY_CSV_FIELDS = [
    "org_id", "org_name", "network_id", "network_name",
    "ssid_name", "ssid_number",
    "tot_checks", "tot_pass", "tot_fail", "tot_not_defined",
    "tot_ignored", "non_standard", "missing_ssid", "overall",
]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Audit SSID configuration against a compliance standard"
    )
    parser.add_argument("--org-id",         required=True,  help="Meraki Organization ID")
    parser.add_argument("--standards-file", required=True,  help="Path to standards Excel file")
    parser.add_argument("--network-id",     default=None,   help="Optional filter: single Network ID")
    parser.add_argument("--csv",            action="store_true", help="Write CSV output")
    parser.add_argument("--xlsx",           action="store_true", help="Write XLSX report")
    args = parser.parse_args()

    standards_path = Path(args.standards_file)
    if not standards_path.exists():
        print(f"ERROR: standards file not found: {standards_path}")
        sys.exit(1)

    print(f"Loading standards from {standards_path.name}...")
    standards      = load_standards(standards_path)
    ignored_fields = load_ignored_fields(standards_path)

    org_key = args.org_id.strip().lower()
    org_standards = {
        k: v for k, v in standards.items() if k[0] == org_key
    }
    if not org_standards:
        print(f"WARNING: no standards defined for org_id '{args.org_id}' in the template.")
        print(f"  Keys found: {[k[0] for k in standards]}")

    print(f"  {len(org_standards)} SSID standard(s) found for this org.")
    for (oid, sname) in org_standards:
        print(f"    standard: ssid_name={sname!r}")
    if ignored_fields:
        print(f"  Ignored fields: {', '.join(sorted(ignored_fields))}")
    print()

    dashboard      = get_dashboard()
    network_lookup = build_network_lookup(dashboard, args.org_id)

    # Resolve org name from networks (best effort)
    org_name = ""
    try:
        orgs = dashboard.organizations.getOrganizations()
        for o in orgs:
            if str(o.get("id")) == str(args.org_id):
                org_name = o.get("name", "")
                break
    except Exception:
        pass

    # Filter to specific network or run all
    if args.network_id:
        if args.network_id not in network_lookup:
            print(f"ERROR: network_id '{args.network_id}' not found in org '{args.org_id}'.")
            print(f"  Available network IDs: {list(network_lookup.keys())[:5]} ...")
            sys.exit(1)
        networks_to_audit = {args.network_id: network_lookup[args.network_id]}
    else:
        networks_to_audit = network_lookup.copy()

    print(f"Auditing {len(networks_to_audit)} network(s) in org '{org_name or args.org_id}'...")
    print()

    all_detail_rows: list[dict] = []

    for network_id, network_name in sorted(networks_to_audit.items(), key=lambda x: x[1]):
        print(f"  [{network_name}]  {network_id}")
        live_ssids = fetch_ssids_for_network(dashboard, network_id)

        if not live_ssids:
            print("    → no SSIDs returned (non-wireless network?), skipping.")
            continue

        detail_rows = audit_network(
            org_id=args.org_id,
            org_name=org_name,
            network_id=network_id,
            network_name=network_name,
            live_ssids=live_ssids,
            standards=standards,
            ignored_fields=ignored_fields,
            single_network_mode=bool(args.network_id),
        )

        # Console summary per network
        pass_count = sum(1 for r in detail_rows if r["result"] == "PASS")
        fail_count = sum(1 for r in detail_rows if r["result"] == "FAIL")
        nd_count   = sum(1 for r in detail_rows if r["result"] in (NOT_DEFINED, "NOT_DEFINED"))
        ign_count  = sum(1 for r in detail_rows if r["result"] == "IGNORED")
        # Count unique rogue/missing SSIDs, not individual field rows
        warn_ssids = {
            r["ssid_name"]
            for r in detail_rows
            if r["result"] in ("NON_STANDARD", "MISSING_SSID")
        }

        print(
            f"    PASS={pass_count}  FAIL={fail_count}  "
            f"NOT_DEFINED={nd_count}  IGNORED={ign_count}  WARNINGS={len(warn_ssids)}"
        )

        # Print ALL check results grouped by SSID
        ssids_seen: list[str] = []
        for r in detail_rows:
            ssid = r["ssid_name"]
            if ssid not in ssids_seen:
                ssids_seen.append(ssid)

        for ssid in ssids_seen:
            ssid_rows = [r for r in detail_rows if r["ssid_name"] == ssid]
            first = ssid_rows[0]
            print(f"\n    SSID #{first['ssid_number']}  {ssid!r}")
            for r in ssid_rows:
                result = r["result"]
                if result == "PASS":
                    icon = "✓"
                elif result == "FAIL":
                    icon = "✗"
                elif result in ("NON_STANDARD", "MISSING_SSID"):
                    icon = "⚠"
                else:
                    icon = "–"

                if result == "MISSING_SSID":
                    print(f"      {icon} {result}")
                elif result == "NON_STANDARD":
                    print(
                        f"      {icon} {result:<14} {r['field']:<28} "
                        f"actual={r['actual']!r}"
                    )
                elif result == "FAIL":
                    print(
                        f"      {icon} {result:<14} {r['field']:<28} "
                        f"expected={r['expected']!r}  actual={r['actual']!r}"
                    )
                elif result in (NOT_DEFINED, "NOT_DEFINED"):
                    print(f"      {icon} NOT_DEFINED    {r['field']:<28} (not audited)")
                elif result == "IGNORED":
                    print(f"      {icon} IGNORED        {r['field']:<28}")
                else:
                    print(f"      {icon} PASS           {r['field']:<28} value={r['actual']!r}")

        all_detail_rows.extend(detail_rows)

    summary_rows = build_summary(all_detail_rows)

    # ── Totals ────────────────────────────────────────────────────────────────
    print()
    tot_pass = sum(r["tot_pass"] for r in summary_rows)
    tot_fail = sum(r["tot_fail"] for r in summary_rows)
    tot_nd   = sum(r["tot_not_defined"] for r in summary_rows)
    tot_warn = sum(1 for r in summary_rows if r["overall"] in ("NON_STANDARD", "MISSING_SSID"))
    print(
        f"TOTAL — PASS={tot_pass}  FAIL={tot_fail}  "
        f"NOT_DEFINED={tot_nd}  WARNINGS={tot_warn}"
    )

    # ── Outputs ───────────────────────────────────────────────────────────────
    suffix = f"_{args.network_id}" if args.network_id else ""

    if args.csv:
        detail_path = OUTPUT_DIR / f"audit_ssid_{args.org_id}{suffix}_detail.csv"
        summary_path = OUTPUT_DIR / f"audit_ssid_{args.org_id}{suffix}_summary.csv"
        write_csv(detail_path,  all_detail_rows, _DETAIL_CSV_FIELDS)
        write_csv(summary_path, summary_rows,    _SUMMARY_CSV_FIELDS)
        print(f"\nCSV detail  → {detail_path}")
        print(f"CSV summary → {summary_path}")

    if args.xlsx:
        xlsx_path = OUTPUT_DIR / f"audit_ssid_{args.org_id}{suffix}.xlsx"
        write_xlsx_report(all_detail_rows, summary_rows, xlsx_path)
        print(f"XLSX report → {xlsx_path}")


if __name__ == "__main__":
    main()
