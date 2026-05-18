from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_DIR
from lib.output import write_csv


def extract_year(date_value: str | None) -> str:
    if not date_value:
        return ""
    value = str(date_value).strip()
    if len(value) >= 4 and value[:4].isdigit():
        return value[:4]
    return ""

def load_org_lookup(org_file: Path) -> dict[str, str]:
    lookup: dict[str, str] = {}

    if not org_file.exists():
        print(f"ERROR: Org File not found, please run list_organizations first {org_file}")
        return lookup

    with org_file.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for row in reader:
            org_id = str(row.get("id","")).strip()
            org_name = str(row.get("name")).strip()

            if org_id:
                lookup[org_id] = org_name
    return lookup

def find_inv_file(output_dir: Path) -> list[Path]:
    files = []
    pattern = re.compile(r"^inventory_(.+)\.csv$")
    for file in output_dir.glob("inventory_*.csv"):
        name = file.name
        if name.startswith("inventory_unmatched_"):
            continue
        if pattern.match(name):
            files.append(file)
    return sorted(files)

def extract_orgid_from_filename(file_path: Path) -> str:
    name = file_path.name
    match = re.match(r"^inventory_([^\.]+)\.csv$", name)
    if not match:
        return ""
    return match.group(1)

def ws_name(name: str, used_names: set[str]) -> str:
    invalid_ch = r'[]:*?/\\'
    clean ="".join("_" if c in invalid_ch else c for c in name)
    clean = clean.strip() or "UNKNOWN_ORG"
    clean = clean[:31]
    orig = clean
    counter = +1

    while clean in used_names:
        suffix = f"_{counter}"
        clean = f"{orig[:31 - len (suffix)]}{suffix}"
        counter =+1
    used_names.add(clean)
    return clean

def style_sheet(ws, table_name: str | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")

    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 35)

    if table_name and ws.max_row > 1 and ws.max_column > 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

def net_in_use(value) -> bool:
    return str(value).strip().lower() in ("true", "yes", "1")

def main() -> None:
    current_year = datetime.now().year
    parser = argparse.ArgumentParser( description="Generate EOL forecast by device model from inventory csv")
    parser.add_argument("--date-field", default="entOfSupportAt", choices=["entOfSupportAt", "endOfSaleAt"], help="Which lifecyle use in forecast")
    parser.add_argument("--start-year", default=current_year, help="Initial Year to Include")
    parser.add_argument("--end-year", default=current_year+7, help="Last Year to Include")
    parser.add_argument("--org-file", default=str(OUTPUT_DIR / "organizations.csv"), help="Path to generated org file")
    parser.add_argument("--only-eol", action="store_true", help="Show only eol")
    parser.add_argument("--output-file", default=str(OUTPUT_DIR / "eol_forecast.xlsx"))
    parser.add_argument("--only-network-associated", action="store_true",help="Exclude devices not associated with network",)

    args = parser.parse_args()


    years = [str(y) for y in range(args.start_year, args.end_year +1)]

    org_lookup = load_org_lookup(Path(args.org_file))
    inventory_files = find_inv_file(OUTPUT_DIR)
    if not inventory_files:
        raise FileNotFoundError(f"No inventory CSV files found in {OUTPUT_DIR}")

    summary: dict[tuple[str, str, str], dict[str, int | str]] = {}
    details_by_org: dict[str, list[dict]] = defaultdict(list)

    for inventory_file in inventory_files:
        org_id = extract_orgid_from_filename(inventory_file)
        org_name = org_lookup.get(org_id, "UNKNOWN ORG")
        print(f"Reading {inventory_file.name} for org {org_id} / {org_name}")

        with inventory_file.open("r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if args.only_network_associated and not net_in_use(row.get("isAssociatedToNetwork")):
                    continue
                model = (row.get("model") or "").strip() or "UNKNOWN"
                lifecycle_year = extract_year(row.get(args.date_field))
                is_eol_range = lifecycle_year in years

                if args.only_eol and not is_eol_range:
                    continue

                key = (org_id, org_name, model)

                if key not in summary:
                    summary[key] = {
                        "orgId": f'="{org_id}"',
                        "Region": org_name,
                        "Device": model,
                        "Tot Number": 0,
                        **{year: 0 for year in years},
                    }


                summary[key]["Tot Number"] +=1


                if lifecycle_year in years:
                    summary[key][lifecycle_year] +=1

                detail_row = dict(row)
                detail_row["orgId"] = org_id
                detail_row["region"] = org_name
                details_by_org[org_name].append(detail_row)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Forecast Summary"

    summary_headers = ["orgId", "Region", "Device", "Tot Number", *years]
    ws_summary.append(summary_headers)


    #output_rows = []

    for key in sorted(summary.keys(), key=lambda x: (x[1], x[2])):
        row = summary[key]
        ws_summary.append([row.get(h, "")for h in summary_headers])

    style_sheet(ws_summary, "ForecastSummary")
    used_sh_name = {"Forecast Summary"}

    # -------------------------
    # GRAND TOTAL ROW
    # -------------------------
    grand_total = {
        "orgId": "TOTAL",
        "Region": "",
        "Device": "",
        "Tot Number": 0,
        **{year: 0 for year in years},
    }

    for row in summary.values():
        grand_total["Tot Number"] += int(row["Tot Number"])
        for year in years:
            grand_total[year] += int(row.get(year, 0))

    ws_summary.append([grand_total.get(h, "") for h in summary_headers])
    last_row = ws_summary.max_row
    for col in range(1, ws_summary.max_column + 1):
        cell = ws_summary.cell(row=last_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF4CC")
        cell.border = Border(top=Side(style="medium"))



    for org_name, rows in sorted(details_by_org.items()):
        if not rows:
            continue

        sheet_name = ws_name(org_name, used_sh_name)
        ws = wb.create_sheet(sheet_name)

        headers = list(rows[0].keys())
        ws.append(headers)

        for row in rows:
            ws.append([row.get(h,"")for h in headers])

        table_name = f"T_{abs(hash(sheet_name)) % 10**8}"
        if not table_name[0].isalpha():
            table_name = f"T_{table_name}"
        style_sheet(ws, table_name)

    output_file = Path(args.output_file)
    wb.save(output_file)
    #output_file = OUTPUT_DIR / f"forecast_{args.date_field}_{args.start_year}_{args.end_year}.csv"
    #fieldnames = ["orgId","Region","Device","Tot Number",*years,]
    #write_csv(output_file, output_rows, fieldnames)


    print(f"\nForecast save in {output_file}")

if __name__ == "__main__":
    main()
