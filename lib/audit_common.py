"""
Shared audit utilities: row building, summary aggregation, XLSX output.
Used by audit_ssid.py, audit_rf_profile.py and audit_ap.py.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from lib.standards import NOT_DEFINED

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
_HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT       = Alignment(horizontal="left",   vertical="center")
_THIN       = Side(style="thin", color="D9D9D9")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_PASS_FILL  = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL  = PatternFill("solid", fgColor="FFC7CE")
_WARN_FILL  = PatternFill("solid", fgColor="FFEB9C")
_ND_FILL    = PatternFill("solid", fgColor="F2F2F2")
_IGN_FILL   = PatternFill("solid", fgColor="DAEEF3")

_RESULT_FILL: dict[str, PatternFill] = {
    "PASS":         _PASS_FILL,
    "FAIL":         _FAIL_FILL,
    "NON_STANDARD": _WARN_FILL,
    "MISSING":      _WARN_FILL,
    "MISSING_SSID": _WARN_FILL,
    NOT_DEFINED:    _ND_FILL,
    "NOT_DEFINED":  _ND_FILL,
    "IGNORED":      _IGN_FILL,
}


# ---------------------------------------------------------------------------
# Row / summary builders
# ---------------------------------------------------------------------------

def make_row(
    org_id: str,
    org_name: str,
    network_id: str,
    network_name: str,
    entity_key: str,        # ssid_name, profile_name, ap_serial, …
    entity_label: str,      # human label for that key (ssid_number, profile_id, …)
    field: str,
    expected: str,
    actual: str,
    result: str,
    standard_level: str = "",
    extra: dict | None = None,
) -> dict:
    row = {
        "org_id":         org_id,
        "org_name":       org_name,
        "network_id":     network_id,
        "network_name":   network_name,
        "entity_key":     entity_key,
        "entity_label":   entity_label,
        "standard_level": standard_level,
        "field":          field,
        "expected":       expected,
        "actual":         actual,
        "result":         result,
    }
    if extra:
        row.update(extra)
    return row


def build_summary(detail_rows: list[dict]) -> list[dict]:
    """One summary row per (network, entity_key)."""
    buckets: dict[tuple, dict] = {}

    for row in detail_rows:
        key = (
            row["org_id"], row["org_name"],
            row["network_id"], row["network_name"],
            row["entity_key"], row["entity_label"],
        )
        if key not in buckets:
            buckets[key] = {
                "org_id":       row["org_id"],
                "org_name":     row["org_name"],
                "network_id":   row["network_id"],
                "network_name": row["network_name"],
                "entity_key":   row["entity_key"],
                "entity_label": row["entity_label"],
                "tot_checks":     0,
                "tot_pass":       0,
                "tot_fail":       0,
                "tot_not_defined":0,
                "tot_ignored":    0,
                "non_standard":   False,
                "missing":        False,
            }

        result = row["result"]
        b = buckets[key]
        if result == "PASS":
            b["tot_checks"] += 1; b["tot_pass"] += 1
        elif result == "FAIL":
            b["tot_checks"] += 1; b["tot_fail"] += 1
        elif result in (NOT_DEFINED, "NOT_DEFINED"):
            b["tot_not_defined"] += 1
        elif result == "IGNORED":
            b["tot_ignored"] += 1
        elif result in ("NON_STANDARD",):
            b["non_standard"] = True
        elif result in ("MISSING", "MISSING_SSID"):
            b["missing"] = True

    rows = []
    for b in buckets.values():
        if b["non_standard"]:           overall = "NON_STANDARD"
        elif b["missing"]:              overall = "MISSING"
        elif b["tot_fail"] > 0:         overall = "FAIL"
        elif b["tot_pass"] > 0:         overall = "PASS"
        else:                           overall = "NOT_DEFINED"
        b["overall"] = overall
        rows.append(b)
    return rows


# ---------------------------------------------------------------------------
# XLSX helpers
# ---------------------------------------------------------------------------

def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
    ws.row_dimensions[1].height = 25


def _add_table(ws, name: str) -> None:
    if ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(tbl)


def _autowidth(ws, min_w=10, max_w=45) -> None:
    for col in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            min(max(length + 2, min_w), max_w)
        )


def _colour_result_cell(ws, row_num: int, result_col: int) -> None:
    cell = ws.cell(row=row_num, column=result_col)
    cell.fill = _RESULT_FILL.get(str(cell.value), _ND_FILL)


def write_xlsx_report(
    detail_rows: list[dict],
    summary_rows: list[dict],
    output_path: Path,
    detail_headers: list[str],
    summary_headers: list[str],
    audit_label: str = "Audit",
) -> None:
    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.freeze_panes = "A2"
    ws_sum.append(summary_headers)
    _style_header(ws_sum)

    for row in sorted(summary_rows,
                      key=lambda r: (r["org_name"], r["network_name"], r["entity_key"])):
        ws_sum.append([row.get(h, "") for h in summary_headers])
        rn = ws_sum.max_row
        _colour_result_cell(ws_sum, rn, summary_headers.index("overall") + 1)
        for cell in ws_sum[rn]:
            cell.border    = _BORDER
            cell.alignment = _LEFT
            cell.font      = Font(name="Arial", size=10)

    _autowidth(ws_sum)
    _add_table(ws_sum, f"{audit_label}Summary")

    # ── Detail sheet ─────────────────────────────────────────────────────────
    ws_det = wb.create_sheet("Detail")
    ws_det.freeze_panes = "A2"
    ws_det.append(detail_headers)
    _style_header(ws_det)

    for row in sorted(detail_rows,
                      key=lambda r: (r["org_name"], r["network_name"],
                                     r["entity_key"], r["field"])):
        ws_det.append([row.get(h, "") for h in detail_headers])
        rn = ws_det.max_row
        _colour_result_cell(ws_det, rn, detail_headers.index("result") + 1)
        for cell in ws_det[rn]:
            cell.border    = _BORDER
            cell.alignment = _LEFT
            cell.font      = Font(name="Arial", size=10)

    _autowidth(ws_det)
    _add_table(ws_det, f"{audit_label}Detail")

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Network lookup helper (shared by all audit tools)
# ---------------------------------------------------------------------------

def build_network_lookup(dashboard, org_id: str) -> dict[str, str]:
    """Return {network_id: network_name} for all networks in the org."""
    networks = dashboard.organizations.getOrganizationNetworks(
        org_id, total_pages="all"
    )
    return {
        str(n["id"]).strip(): n.get("name", "")
        for n in networks
        if isinstance(n, dict) and n.get("id")
    }
