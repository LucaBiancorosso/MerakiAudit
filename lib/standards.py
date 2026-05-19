from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


# Fields that are identity keys, not audited config fields
_KEY_FIELDS = {"org_id", "network_id", "ssid_name"}

# Sentinel for a blank standard cell
NOT_DEFINED = "NOT_DEFINED"


def _cell_to_str(value: Any) -> str:
    """
    Robustly convert an Excel cell value to a clean string, handling all
    the ways Excel and openpyxl can mangle data:

      - float with no fractional part (e.g. 123456.0 -> '123456')
        Happens when a numeric org_id/network_id column is not formatted as Text.
      - bool stored as Python bool (e.g. True -> 'True', False -> 'False')
        Happens when the user types TRUE/FALSE and Excel stores as boolean.
      - scientific notation floats (e.g. 1.23457e+11)
        Happens for large IDs that Excel auto-converts to scientific notation.
      - plain int / str: converted with str() as usual.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        # Must come before int check — bool is a subclass of int in Python
        return str(value)           # True -> 'True', False -> 'False'
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))  # 123456.0 -> '123456'
        return str(value)
    return str(value).strip()


def _normalise(value: Any) -> str:
    """Return a clean lowercase string for comparison."""
    return _cell_to_str(value).strip().lower()


def load_standards(path: Path) -> dict[tuple[str, str, str], dict[str, Any]]:
    """
    Read the Standards sheet and return a lookup keyed by
    (org_id, network_id, ssid_name) — all lowercase.

    network_id is optional in the sheet: leave the cell blank to define an
    org-wide standard.  A network-specific row (network_id filled in) takes
    priority over an org-wide row during lookup; see resolve_standard().

    Each value is a dict of {field: expected_value}.
    Blank cells are stored as NOT_DEFINED sentinel so callers can distinguish
    'not configured' from 'empty string expected'.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Standards" not in wb.sheetnames:
        raise ValueError(f"'Standards' sheet not found in {path}")

    ws = wb["Standards"]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Standards sheet is empty")

    headers = [_cell_to_str(h).strip() for h in rows[0]]

    standards: dict[tuple[str, str, str], dict[str, Any]] = {}

    for row in rows[1:]:
        row_dict = dict(zip(headers, row))

        org_id     = _normalise(row_dict.get("org_id"))
        network_id = _normalise(row_dict.get("network_id"))   # "" when blank → org-level
        ssid_name  = _cell_to_str(row_dict.get("ssid_name")).strip()

        # Skip blank rows or the hint row in the template
        if not org_id or not ssid_name or ssid_name.lower() in ("ssid_name", "← required"):
            continue

        key = (org_id, network_id, ssid_name.lower())
        fields: dict[str, Any] = {}

        for header, value in row_dict.items():
            if header in _KEY_FIELDS or not header:
                continue
            cell_str = _cell_to_str(value).strip()
            fields[header] = NOT_DEFINED if cell_str == "" else cell_str

        standards[key] = fields

    return standards


def resolve_standard(
    standards: dict[tuple[str, str, str], dict[str, Any]],
    org_id: str,
    network_id: str,
    ssid_name: str,
) -> dict[str, Any] | None:
    """
    Look up the most specific standard for an SSID, with fallback:
      1. Network-specific:  (org_id, network_id, ssid_name)
      2. Org-wide:          (org_id, "",          ssid_name)
      3. None → SSID is not in the standard at all (NON_STANDARD)
    """
    oid   = org_id.strip().lower()
    nid   = network_id.strip().lower()
    sname = ssid_name.strip().lower()

    return (
        standards.get((oid, nid,  sname))   # network-specific first
        or standards.get((oid, "", sname))   # org-wide fallback
    )


def load_ignored_fields(path: Path) -> set[str]:
    """
    Read the Ignore sheet and return a set of field names to skip globally.
    Returns empty set if sheet is missing or empty.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Ignore" not in wb.sheetnames:
        return set()

    ws = wb["Ignore"]
    ignored: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_str = _cell_to_str(row[0]).strip()
        if cell_str:
            ignored.add(cell_str)

    return ignored


def compare_field(
    field: str,
    expected: str,
    actual: Any,
    ignored_fields: set[str],
) -> dict[str, str]:
    """
    Compare a single field and return a result dict with keys:
        field, expected, actual, result
    """
    if field in ignored_fields:
        return {
            "field":    field,
            "expected": expected,
            "actual":   _cell_to_str(actual),
            "result":   "IGNORED",
        }

    if expected == NOT_DEFINED:
        return {
            "field":    field,
            "expected": "",
            "actual":   _cell_to_str(actual),
            "result":   NOT_DEFINED,
        }

    actual_str   = _normalise(actual)
    expected_str = _normalise(expected)

    # Special handling for radiusHosts: compare as unordered comma-separated sets
    if field == "radiusHosts":
        actual_set   = {h.strip() for h in actual_str.split(",") if h.strip()}
        expected_set = {h.strip() for h in expected_str.split(",") if h.strip()}
        result = "PASS" if actual_set == expected_set else "FAIL"
    else:
        result = "PASS" if actual_str == expected_str else "FAIL"

    return {
        "field":    field,
        "expected": expected,
        "actual":   _cell_to_str(actual),
        "result":   result,
    }


# ===========================================================================
# RF Profile standards
# ===========================================================================

_RF_KEY_FIELDS = {"org_id", "network_id", "profile_name"}


def load_rf_standards(path: Path) -> dict[tuple[str, str, str], dict[str, Any]]:
    """
    Read the RFProfiles sheet.
    Key: (org_id, network_id, profile_name) — all lowercase.
    network_id is optional (blank = org-wide default).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if "RFProfiles" not in wb.sheetnames:
        return {}

    ws = wb["RFProfiles"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [_cell_to_str(h).strip() for h in rows[0]]
    standards: dict[tuple[str, str, str], dict[str, Any]] = {}

    for row in rows[1:]:
        rd = dict(zip(headers, row))
        org_id      = _normalise(rd.get("org_id"))
        network_id  = _normalise(rd.get("network_id"))
        profile_name = _cell_to_str(rd.get("profile_name")).strip()

        if not org_id or not profile_name or profile_name.lower() in ("profile_name", "← required"):
            continue

        key = (org_id, network_id, profile_name.lower())
        fields: dict[str, Any] = {}
        for header, value in rd.items():
            if header in _RF_KEY_FIELDS or not header:
                continue
            cell_str = _cell_to_str(value).strip()
            fields[header] = NOT_DEFINED if cell_str == "" else cell_str

        standards[key] = fields

    return standards


def resolve_rf_standard(
    standards: dict[tuple[str, str, str], dict[str, Any]],
    org_id: str,
    network_id: str,
    profile_name: str,
) -> dict[str, Any] | None:
    oid   = org_id.strip().lower()
    nid   = network_id.strip().lower()
    pname = profile_name.strip().lower()
    return (
        standards.get((oid, nid,  pname))
        or standards.get((oid, "", pname))
    )


# ===========================================================================
# AP Config standards
# ===========================================================================

_AP_KEY_FIELDS = {"org_id", "network_id"}


def load_ap_standards(path: Path) -> dict[tuple[str, str], dict[str, Any]]:
    """
    Read the APConfig sheet.
    Key: (org_id, network_id) — network_id blank = org-wide default.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if "APConfig" not in wb.sheetnames:
        return {}

    ws = wb["APConfig"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [_cell_to_str(h).strip() for h in rows[0]]
    standards: dict[tuple[str, str], dict[str, Any]] = {}

    for row in rows[1:]:
        rd = dict(zip(headers, row))
        org_id     = _normalise(rd.get("org_id"))
        network_id = _normalise(rd.get("network_id"))

        if not org_id or org_id in ("org_id", "← required"):
            continue

        key = (org_id, network_id)
        fields: dict[str, Any] = {}
        for header, value in rd.items():
            if header in _AP_KEY_FIELDS or not header:
                continue
            cell_str = _cell_to_str(value).strip()
            fields[header] = NOT_DEFINED if cell_str == "" else cell_str

        standards[key] = fields

    return standards


def resolve_ap_standard(
    standards: dict[tuple[str, str], dict[str, Any]],
    org_id: str,
    network_id: str,
) -> dict[str, Any] | None:
    oid = org_id.strip().lower()
    nid = network_id.strip().lower()
    return (
        standards.get((oid, nid))
        or standards.get((oid, ""))
    )


# ===========================================================================
# Subnet check utility (for AP management IP audit)
# ===========================================================================

import ipaddress as _ipaddress


def ip_in_allowed_subnets(ip: str, subnets_csv: str) -> bool:
    """
    Return True if `ip` falls within any of the comma-separated CIDR subnets.
    Gracefully ignores blank or malformed entries.
    """
    if not ip or not subnets_csv:
        return False
    try:
        addr = _ipaddress.ip_address(ip.strip())
    except ValueError:
        return False

    for subnet_str in subnets_csv.split(","):
        subnet_str = subnet_str.strip()
        if not subnet_str:
            continue
        try:
            net = _ipaddress.ip_network(subnet_str, strict=False)
            if addr in net:
                return True
        except ValueError:
            continue
    return False


# ===========================================================================
# Switch Interface standards
# ===========================================================================

_SW_KEY_FIELDS = {"org_id", "network_id", "role"}


def expand_vlan_ranges(vlan_str: str) -> set[int]:
    """
    Expand a Meraki VLAN string like '1,3,5-10,20' into a set of ints.
    'all' expands to the full 1-4094 range.
    Gracefully skips malformed tokens.
    """
    if not vlan_str:
        return set()
    s = str(vlan_str).strip().lower()
    if s == "all":
        return set(range(1, 4095))
    result: set[int] = set()
    for token in s.split(","):
        token = token.strip()
        if not token:
            continue
        if "-" in token:
            parts = token.split("-", 1)
            try:
                lo, hi = int(parts[0]), int(parts[1])
                result.update(range(lo, hi + 1))
            except ValueError:
                continue
        else:
            try:
                result.add(int(token))
            except ValueError:
                continue
    return result


def load_switch_standards(path: Path) -> dict[tuple[str, str, str], dict[str, Any]]:
    """
    Read the SwitchInterfaces sheet.
    Key: (org_id, network_id, role) — all lowercase.
    network_id blank = org-wide default.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if "SwitchInterfaces" not in wb.sheetnames:
        return {}

    ws = wb["SwitchInterfaces"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [_cell_to_str(h).strip() for h in rows[0]]
    standards: dict[tuple[str, str, str], dict[str, Any]] = {}

    for row in rows[1:]:
        rd = dict(zip(headers, row))
        org_id     = _normalise(rd.get("org_id"))
        network_id = _normalise(rd.get("network_id"))
        role       = _cell_to_str(rd.get("role")).strip().lower()

        if not org_id or not role or role in ("role", "← required"):
            continue

        key = (org_id, network_id, role)
        fields: dict[str, Any] = {}
        for header, value in rd.items():
            if header in _SW_KEY_FIELDS or not header:
                continue
            cell_str = _cell_to_str(value).strip()
            fields[header] = NOT_DEFINED if cell_str == "" else cell_str

        standards[key] = fields

    return standards


def resolve_switch_standard(
    standards: dict[tuple[str, str, str], dict[str, Any]],
    org_id: str,
    network_id: str,
    role: str,
) -> dict[str, Any] | None:
    oid  = org_id.strip().lower()
    nid  = network_id.strip().lower()
    role = role.strip().lower()
    return (
        standards.get((oid, nid,  role))
        or standards.get((oid, "", role))
    )


# ===========================================================================
# VLAN → role mapping
# ===========================================================================

_VLAN_KEY_FIELDS = {"org_id", "network_id", "vlan_id"}


def load_vlan_roles(path: Path) -> dict[tuple[str, str, str], str]:
    """
    Read the VlanRoles sheet.
    Key: (org_id, network_id, vlan_id_str) — all lowercase.
    Returns {key: role_name}.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if "VlanRoles" not in wb.sheetnames:
        return {}

    ws = wb["VlanRoles"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [_cell_to_str(h).strip() for h in rows[0]]
    result: dict[tuple[str, str, str], str] = {}

    for row in rows[1:]:
        rd = dict(zip(headers, row))
        org_id     = _normalise(rd.get("org_id"))
        network_id = _normalise(rd.get("network_id"))
        vlan_id    = _cell_to_str(rd.get("vlan_id")).strip()
        role       = _cell_to_str(rd.get("role")).strip().lower()

        if not org_id or not vlan_id or not role:
            continue

        result[(org_id, network_id, vlan_id)] = role

    return result


def resolve_vlan_role(
    vlan_roles: dict[tuple[str, str, str], str],
    org_id: str,
    network_id: str,
    vlan_id: int | str,
) -> str | None:
    """
    Look up role for a VLAN, network-specific first then org-wide.
    Returns role string or None if not mapped.
    """
    oid  = org_id.strip().lower()
    nid  = network_id.strip().lower()
    vid  = str(vlan_id).strip()
    return (
        vlan_roles.get((oid, nid, vid))
        or vlan_roles.get((oid, "", vid))
    )
